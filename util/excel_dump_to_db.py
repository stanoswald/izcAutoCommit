import configparser

from openpyxl import load_workbook
import pymysql


def get_from_excel(beg, end):
    data = list()
    wb = load_workbook('信息收集表（收集结果）.xlsx')
    ws = wb['信息收集表']
    for row in ws[beg:end]:
        sno = int(row[2].value)
        name = row[3].value
        province, city, district = row[4].value.split("/")
        data.append((sno, name, province, city, district))

    print("excel读取完毕")
    return data


def dump_to_db(data: list[tuple]):
    conf = configparser.ConfigParser()
    conf.read("db.ini", encoding="utf-8")
    host = conf['mysql']['host']
    usr = conf['mysql']['username']
    pwd = conf['mysql']['password']
    db = conf['mysql']['db']

    conn = pymysql.connect(host=host, user=usr, password=pwd, database=db, charset='utf8')
    cur = conn.cursor()
    sql = "INSERT INTO tbl_stu(sno, name, province, city, district) VALUES (%s,%s,%s,%s,%s)"
    cur.executemany(sql, data)

    conn.commit()
    print("数据库写入成功")
    conn.close()


if __name__ == '__main__':
    beg = int(input("请输入开始行号："))
    end = int(input("请输入结束行号："))
    stus = get_from_excel(beg, end)
    dump_to_db(stus)
